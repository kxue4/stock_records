#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
股票收盘价记录脚本
支持 A股、港股、美股，将每日收盘价写入 Excel。
Excel 格式：行 = 日期，列 = 股票代码，单元格 = 收盘价。
依赖：pip install yfinance openpyxl
"""

import os
from datetime import datetime, timedelta

import openpyxl
import yfinance as yf

# ======================== 在此处填写股票代码 ========================
# A股：上交所加 .SS 后缀，深交所加 .SZ 后缀
# 港股：加 .HK 后缀
# 美股：直接填写代码，无需后缀
STOCK_CODES: list[str] = [
    # ---------- A股 ----------
    "000001.SS",  #上证指数
    "000300.SS",  # 沪深300 
    "600519.SS",  # 贵州茅台
    # ---------- 港股 ----------
    "0700.HK",    # 腾讯控股
    "0883.HK",    # 中国海洋石油
    # ---------- 美股 ----------
    "^DJI",        # 道琼斯
    "PDD",        # 拼多多
]

# ======================== 在此处填写查询时间范围 ========================
# 格式：YYYY-MM-DD
START_DATE = "2026-01-01"
END_DATE = "2026-03-13"

# ======================== Excel 文件路径 ========================
EXCEL_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "stock_records.xlsx")


def load_or_create_workbook(filepath: str, stock_codes: list[str]) -> openpyxl.Workbook:
    """加载已有 Excel，若不存在则新建并写入表头（A1 为 '日期'，后续列为股票代码）。"""
    if os.path.exists(filepath):
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        # 检查是否需要追加新的股票代码列
        existing_codes = []
        for col in range(2, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val:
                existing_codes.append(str(val))
        for code in stock_codes:
            if code not in existing_codes:
                new_col = ws.max_column + 1
                ws.cell(row=1, column=new_col, value=code)
        return wb

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "股票记录"
    ws.cell(row=1, column=1, value="日期")
    for i, code in enumerate(stock_codes):
        ws.cell(row=1, column=i + 2, value=code)
    return wb


def get_code_col_map(ws) -> dict[str, int]:
    """构建 {股票代码: 列号} 的映射。"""
    mapping = {}
    for col in range(2, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            mapping[str(val)] = col
    return mapping


def get_date_row_map(ws) -> dict[str, int]:
    """构建 {日期字符串: 行号} 的映射。"""
    mapping = {}
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val:
            mapping[str(val).split(" ")[0]] = row
    return mapping


def fetch_and_write(stock_codes: list[str], start_date: str, end_date: str, filepath: str) -> None:
    """获取指定时间范围内的收盘价并写入 Excel。
    停牌或休市的日期记录为 '/'。"""
    wb = load_or_create_workbook(filepath, stock_codes)
    ws = wb.active
    code_col = get_code_col_map(ws)
    date_row = get_date_row_map(ws)

    # 第一步：生成时间范围内的所有连续日期
    start_dt = datetime.strptime(start_date, "%Y-%m-%d")
    end_dt = datetime.strptime(end_date, "%Y-%m-%d")
    all_dates: list[str] = []
    current = start_dt
    while current < end_dt:
        all_dates.append(current.strftime("%Y-%m-%d"))
        current += timedelta(days=1)

    # 第二步：获取所有股票数据
    # all_data: {code: {date_str: close_price}}
    all_data: dict[str, dict[str, float]] = {}

    for code in stock_codes:
        print(f"正在获取 {code} 的数据...")
        ticker = yf.Ticker(code)
        hist = ticker.history(start=start_date, end=end_date)

        if hist.empty:
            print(f"  警告：{code} 未获取到数据，请检查代码是否正确。")
            all_data[code] = {}
            continue

        code_data = {}
        for date, row_data in hist.iterrows():
            date_str = date.strftime("%Y-%m-%d")
            code_data[date_str] = round(row_data["Close"], 2)
        all_data[code] = code_data

    # 第三步：确保所有日期都有对应行
    for date_str in all_dates:
        if date_str not in date_row:
            new_row = ws.max_row + 1
            ws.cell(row=new_row, column=1, value=date_str)
            date_row[date_str] = new_row

    # 第四步：写入数据，无数据的日期填 "/"
    new_count = 0
    skip_count = 0

    for code in stock_codes:
        col = code_col[code]
        code_data = all_data.get(code, {})

        for date_str in all_dates:
            target_row = date_row[date_str]

            # 已有数据则跳过
            if ws.cell(row=target_row, column=col).value is not None:
                skip_count += 1
                continue

            if date_str in code_data:
                ws.cell(row=target_row, column=col, value=code_data[date_str])
            else:
                # 该股票在此日期无数据（停牌或该市场休市）
                ws.cell(row=target_row, column=col, value="")
            new_count += 1

    # 按日期排序（表头行除外）
    data_rows = []
    for row in range(2, ws.max_row + 1):
        data_rows.append([ws.cell(row=row, column=c).value for c in range(1, ws.max_column + 1)])
    data_rows.sort(key=lambda r: str(r[0]) if r[0] else "")
    for i, row_data in enumerate(data_rows):
        for j, val in enumerate(row_data):
            ws.cell(row=i + 2, column=j + 1, value=val)

    wb.save(filepath)
    print(f"\n完成！新写入 {new_count} 条，跳过已有 {skip_count} 条。")
    print(f"文件保存至：{filepath}")


if __name__ == "__main__":
    fetch_and_write(STOCK_CODES, START_DATE, END_DATE, EXCEL_FILE)
